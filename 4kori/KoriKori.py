# -*- coding: utf-8 -*-

import unicodedata #맥OS의 풀어쓰기를 해결하기 위한 라이브러리
import time #sleep 을 쓰려고 

from bs4 import BeautifulSoup #크롤링을 위해
from selenium import webdriver #안보이는 크롬브라우저를 쓰기 위해
from selenium.webdriver.common.keys import Keys #카테고리 클릭을 위해
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook #엑셀파일을 만들기 위해
from datetime import datetime #오늘 날짜와 시간 받아오려고
## pip install selenium
## pip install openpyxl
## pip install beautifulsoup


##크롬드라이버 구동을 위해 최소한으로 주는 대기시간(초)
delay = 3

##txt 파일을 하나 만듭니다. 이미 있으면 내용 다지움. f 라는 이름을 붙임. 닫는 것 잊지 말 것.
f = open("kori.txt",'w', encoding='utf8')
print("시작") #콘솔창에 현재 뭐 하는 지 출력
f.write(datetime.today().strftime("%Y/%m/%d %H:%M:%S")+"\n") #오늘 날짜와 시간 적고 줄바꿈

##크롬드라이버를 Headless모드(창없음)로 동작하게 하기 위한 옵션 구성
myoptions = webdriver.ChromeOptions() #기본옵션
myoptions.add_argument('--headless') #안보임
myoptions.add_argument('--window-size=1920,1080') #라바나가 크기 작으면 카테고리 버튼이 안보여서 키움
#myoptions.add_argument("--disable-gpu") #그래픽이 후진 컴에서도 되게
myoptions.add_experimental_option('excludeSwitches', ['enable-logging']) #콘솔창에 한 줄 나오는 거 숨김
## 혹은 myoptions.add_argument("--disable-gpu")

f.write(str(myoptions.arguments)+"\n") #옵션 뭐 했는 지 기록
driver = webdriver.Chrome('chromedriver.exe', options=myoptions) #driver 라는 이름으로 크롬브라우저를 사용
driver.implicitly_wait(delay)

## 암묵적으로 웹 자원 로드를 위해 3초까지 기다려 준다.
print(str(delay)+"초 대기"+"\n")

## 크롬드라이버로 url의 내용을 가져온다.
myurl = "https://lavana.ac"
driver.get(myurl)

#print(myurl)
f.write(myurl+"의 카테고리 분석"+"\n")

#CATEGORY 버튼을 클릭해서 펼친다.
driver.find_element_by_xpath('//*[@id="root"]/div/div/header/div/div[1]/button').send_keys(Keys.ENTER) 
#위의 내용은 크롬브라우저 F12에서 elements에 나오는 놈들에서 찾아서 우클릭 copy > copy selector 의 내용을 약간 가공

html = driver.page_source #카테고리 내용 펼쳐진 채로 웹페이지 소스를 html 이름에 저장
soup = BeautifulSoup(html, 'html.parser') #크롤링 라이브러리를 사용하여 soup에 내용을 저장, 탐색하여 추출할 수 있음
print(myurl + "\n")

my_title = soup.select('#category-menu-list-grow > ul > a') #soup에 들어있는 것 중 해당 seletor 부분을 my_title에 저장
#print(my_title)

wb = Workbook() #wb 라는 이름으로 엑셀파일 하나 생성
ws = wb.active #ws 라는 이름으로 엑셀파일을 활성화
print("엑셀파일 생성합니다."+"\n")
ws['A1']='Hello Kori' #테스트삼아 A1셀에 글자 넣음

i=0 #1차 for 루프를 돌리기 위한 인덱스 변수
cnt2=0 #카운트의 합계값을 저장하려는 변수
driver2 = webdriver.Chrome('chromedriver.exe', options=myoptions) #driver2라는 이름으로 크롬브라우저 한 개 더 쓸거임. 옵션은아까맨치로
#WebDriverWait(driver2, 3)
driver2.implicitly_wait(delay+3)

for p in my_title:
    i=i+1 #엑셀은 1부터 시작하니까 일단 1 올리고 시작
    
    cattitle = p.text.strip() #p라는 1단위의 데이터에 대해 태그 사이의 내용을 저장. 카테고리 이름
    ws.cell(row=i+1, column=1).value = cattitle #그걸 엑셀에 기입하는데 1칸에는 아까 적은 헬로 코리 있으니 그다음 칸을 지정하는 게 i+1
    
    caturl = myurl + p.get('href') #p라는 1단위의 데이터에 대해 태그 내부의 href 속성값을 가져옴. html 소스를 보면서 수정할 것. 카테고리 주소
    ws.cell(row=i+1, column=2).value = caturl #그걸 엑셀에 기입
    
    print(cattitle + " : "+caturl) #콘솔창에 출력
    print("---------------------------------------------------")
    f.write(" "+"\n") #한줄 띄우고
    f.write(cattitle + " 에는 다음의 클래스들이 등록되어 있습니다."+"\n") #txt 파일에 기입
    f.write("---------------------------------------------------"+"\n")
    
    driver2.get(caturl) #카테고리 주소 페이지를 접속
    driver2.implicitly_wait(delay+3)
    time.sleep(3)

    try:
        #element1 = WebDriverWait(driver2, 45).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "#main > div > section"))) #"#main > div > section"
        #element2 = WebDriverWait(driver2, 45).until(EC.presence_of_element_located((By.ID, 'main'))) 
        
        html2 = driver2.page_source #html2에 읽은 거 넣고
        soup2 = BeautifulSoup(html2, 'html.parser') #soup2에 분석한 내용 저장.
        
        #print(soup2) #디버그용
       
        #lecturesincat = soup2.select('#main > div > section.undefined.jss287 > div > a') #seletor 찾느라..
        #lecturesincat = soup2.select('section > div > a > div.MuiCardContent-root > p.MuiTypography-root') #soup2 내용에서 클래스가 나열된 곳을 추출 
        lecturesincat = soup2.select('main > div > section > div > a > div > p') #soup2 내용에서 클래스가 나열된 곳을 추출
        

        # main > div > section > div.title 이 있으면 진행하고 없으면 재시도...? 3번?

    
        #lecturesincat = soup2.select('section > div > a > div > p') #soup2 내용에서 클래스가 나열된 곳을 추출 
    
        #main > div > section.undefined.jss1467 > div:nth-child(1) > a > div.MuiCardContent-root.jss1482 > p.MuiTypography-root.jss1483.MuiTypography-subtitle2
        #main > div > section.undefined.jss1467 > div:nth-child(11) > a > div.MuiCardContent-root.jss1482 > p.MuiTypography-root.jss1483.MuiTypography-subtitle2
    
        #print(lecturesincat) #디버그용

    finally:
        
        cnt = 0 #클래스 갯수를 세어 저장할 변수
        for n in lecturesincat:
            n2 = unicodedata.normalize('NFC',n.text.strip()) #n의 내용 중 풀어쓰기 된 거 있으면 다시 모아쓰기해서 n2에 저장
            print(n2)
            try:
                f.write(n2 + "\n") #파일에 저장할 때 윈도의 cp949 인코딩이면 파이썬 기본 인코딩인 utf-8과 달라서 인코딩이 다른 두 문자열을 합칠 수 없다는 에러날거임
            except:
                print("err")
                n3 = n2.decode('cp949') #일단 cp949를 유니코드로 디코딩해서 n3에 넣고
                n4 = unicodedata.normalize('NFC',n3) #유니코드상태에서 풀어쓰기 된 거 있으면 다시 모아쓰기 해서 n4에 저장
                n5 = n4.encode('utf-8') #다시 유니코드를 utf-8로 인코딩해서 n5에 넣고
                f.write(n5+"\n") #파일에 기록

            cnt = cnt + 1 #클래스 갯수 한 개 카운트
            #다시 for n 부분으로 돌아감

        #print(caturl + " 의 내용 끝입니다.")

        print("---------------------------------------------------")
        print(cattitle + " 에 등록된 클래스는 모두 " + str(cnt/2) + "개 입니다.") #클래스 갯수 세는 for 루프 다 돌고 나서 카운트 변수를 출력
        print(" "+"\n")
        f.write("---------------------------------------------------"+"\n")
        f.write(cattitle + " 에 등록된 클래스는 모두 " + str(cnt/2) + "개 입니다."+"\n")
        f.write(" "+"\n")
        ws.cell(row=i+1, column=3).value = cnt/2 #카운트 변수를 엑셀에도 기록

        cnt2=cnt2+(cnt/2) #합계 변수에 방금 센 카운트 변수를 더해 줌
        cnt=0 #카운트 변수 다시 0으로
        soup2.clear #분석 내용2 다 지우고    
        #다시 for p 부분으로 돌아감
    

driver2.close()
driver2.quit()

print(myurl + " 의 분석이 끝났습니다.")
#driver2.quit() #크롬브라우저2 닫고

ws.cell(row=i+2, column=3).value = "=sum(C2:C"+str(i+1)+")" #for p 다 돌고 나서 i+1 칸을 마지막으로 기록했을 테니 그 아랫칸에 총합계 넣을 거임. i+2번째 칸. 내용은 엑셀함수 내용 씀.
#문제는 파이썬 엑셀 라이브러리는 엑셀함수내용을 기록해봤자 엑셀파일을 진짜 엑셀에서 열 때까지 sum함수를 계산이 돌아가진 않는 게 문제. 그래서 cnt2 합계 변수를 쓴 것.
wb.save('kori.xlsx') #엑셀 파일 저장
print("kori.xlsx 엑셀파일 저장했습니다.")
print("카테고리는 모두 "+str(i)+"개이며 등록된 클래스의 수는 중복을 포함하여 모두 "+str(cnt2)+"개입니다.") #cnt2 변수내용 써먹음
f.write("카테고리는 모두 "+str(i)+"개이며 등록된 클래스의 수는 중복을 포함하여 모두 "+str(cnt2)+"개입니다."+"\n")

driver.quit() #크롬브라우저1 닫고
soup.clear #분석 내용1 지우고
print("끝")
f.close() #txt 파일 저장하고 닫고